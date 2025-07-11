import os
import sys
import re
import smtplib
import requests
import pandas as pd
from datetime import datetime
from collections import defaultdict
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from tabulate import tabulate
from openpyxl import load_workbook

# === CONFIGURABLE RECIPIENTS ===
EMAIL_RECIPIENTS = [
    "shivammusterya@gmail.com",
    "musteryashivam@gmail.com",
    "shivam.m@browserstack.com"
]

EMAIL_CC = [
    "shivam.m@browserstack.com"
]

class JenkinsLogParser:
    def __init__(self, base_job_url, username, api_token):
        self.base_job_url = base_job_url.rstrip("/")
        self.auth = (username, api_token)

    def get_build_info(self, build_number):
        url = f"{self.base_job_url}/{build_number}/api/json"
        response = requests.get(url, auth=self.auth)
        return response.json() if response.status_code == 200 else None

    def fetch_console_log(self, build_number):
        url = f"{self.base_job_url}/{build_number}/consoleText"
        response = requests.get(url, auth=self.auth)
        return response.text if response.status_code == 200 else None

    def parse_counts_from_line(self, line):
        total_match = re.match(r"\d+", line)
        total = int(total_match.group(0)) if total_match else 0

        failed = passed = skipped = 0
        counts_match = re.search(r"\((.*?)\)", line)
        if counts_match:
            for count, label in re.findall(r"(\d+) (\w+)", counts_match.group(1)):
                count = int(count)
                label = label.lower()
                if label == "failed": failed = count
                elif label == "passed": passed = count
                elif label == "skipped": skipped = count

        return {"total": total, "failed": failed, "passed": passed, "skipped": skipped}

    def extract_summary_counts(self, console_output):
        pattern = re.compile(r"^(?:\[(\d{4}-\d{2}-\d{2}.*?)Z\]|\b(\d{2}:\d{2}:\d{2}))\s+(\d+ \w+ \(.*\))", re.MULTILINE)
        matches = pattern.findall(console_output)
        if len(matches) >= 2:
            return {
                "scenarios": self.parse_counts_from_line(matches[0][2]),
                "steps": self.parse_counts_from_line(matches[1][2])
            }
        return {}

    def extract_failures(self, console_output):
        match = re.search(r"Failures:(.*?)(?=\d+m\d+\.\d+s \(executing steps: )", console_output, re.DOTALL)
        if match:
            return re.findall(r"Scenario: (.*?)\s*#", match.group(1))
        return []

    def extract_env(self, output):
        patterns = [
            r"Started by timer with parameters: \{[^}]*ENV=([^,}]+)",
            r"ENV=([A-Za-z0-9_]+)",
            r"Environment[:=]\s*([A-Za-z0-9_]+)",
            r"Run environment:\s*([A-Za-z0-9_]+)"
        ]
        for pat in patterns:
            match = re.search(pat, output, re.IGNORECASE)
            if match:
                return match.group(1).strip().lower()
        return "unknown"

    def aggregate_last_n_builds_by_date_and_env(self, n=5):
        url = f"{self.base_job_url}/lastBuild/api/json"
        resp = requests.get(url, auth=self.auth)
        if resp.status_code != 200:
            print("❌ Failed to get latest build info.")
            return {}

        latest = resp.json()["number"]
        data = defaultdict(lambda: defaultdict(lambda: {
            "builds": [],
            "scenarios": {"total": 0, "passed": 0, "failed": 0, "skipped": 0},
            "steps": {"total": 0, "passed": 0, "failed": 0, "skipped": 0},
            "failed_scenarios": []
        }))

        for i in range(latest, latest - n, -1):
            info = self.get_build_info(i)
            if not info: continue

            date_str = datetime.fromtimestamp(info["timestamp"] // 1000).strftime('%Y-%m-%d')
            console = self.fetch_console_log(i)
            if not console: continue

            env = self.extract_env(console)
            if env not in ["prod", "preprod"]:
                env = "ran_manually"

            counts = self.extract_summary_counts(console)
            fails = self.extract_failures(console)
            entry = data[date_str][env]

            entry["builds"].append(i)
            for key in ["scenarios", "steps"]:
                for typ in ["total", "passed", "failed", "skipped"]:
                    entry[key][typ] += counts.get(key, {}).get(typ, 0)
            entry["failed_scenarios"].extend(fails)

        return dict(data)

def save_to_excel(df, filename="jenkins_summary.xlsx", sheet="Summary"):
    if os.path.exists(filename):
        book = load_workbook(filename)
        if sheet in book.sheetnames:
            df_old = pd.read_excel(filename, sheet_name=sheet)
            df = pd.concat([df_old, df], ignore_index=True)
            df.drop_duplicates(subset=["date", "builds", "environment"], keep="last", inplace=True)
    with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
    print(f"✅ Excel saved: {filename}")

def colorize_stability(value):
    if value >= 95:
        return f'<td style="color:green;font-weight:bold">{value:.2f}</td>'
    elif value >= 80:
        return f'<td style="color:orange;font-weight:bold">{value:.2f}</td>'
    else:
        return f'<td style="color:red;font-weight:bold">{value:.2f}</td>'

def send_email_report(subject, df, attachments):
    sender = os.getenv("EMAIL_USER")
    password = os.getenv("EMAIL_PASS")

    if not sender or not password:
        print("❌ Missing EMAIL_USER or EMAIL_PASS env vars.")
        return

    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = ", ".join(EMAIL_RECIPIENTS)
    msg["Cc"] = ", ".join(EMAIL_CC)
    msg["Subject"] = subject

    # HTML Table with highlighted stability
    headers = df.columns.tolist()
    rows = []
    for _, row in df.iterrows():
        html_row = ""
        for col in headers:
            if col == "stability":
                html_row += colorize_stability(row[col])
            else:
                html_row += f"<td>{row[col]}</td>"
        rows.append(f"<tr>{html_row}</tr>")

    html_table = f"""
    <html><body>
    <h2>📊 Jenkins Daily Report</h2>
    <table border="1" cellpadding="6" cellspacing="0">
    <thead><tr>{"".join([f"<th>{h}</th>" for h in headers])}</tr></thead>
    <tbody>{"".join(rows)}</tbody>
    </table>
    </body></html>
    """

    msg.attach(MIMEText(html_table, "html"))

    for path in attachments:
        with open(path, "rb") as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(path))
            part["Content-Disposition"] = f'attachment; filename="{os.path.basename(path)}"'
            msg.attach(part)

    recipients = EMAIL_RECIPIENTS + EMAIL_CC
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, password)
            server.sendmail(sender, recipients, msg.as_string())
        print("📧 Email sent successfully.")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")

if __name__ == "__main__":
    base_url = "https://minion.browserstack.com/job/QA/job/LcncAPITests"
    username = os.getenv("JENKINS_USER")
    token = os.getenv("JENKINS_TOKEN")

    if not username or not token:
        print("❌ Missing Jenkins credentials.")
        sys.exit(1)

    parser = JenkinsLogParser(base_url, username, token)
    results = parser.aggregate_last_n_builds_by_date_and_env(n=10)

    if not results:
        print("❌ No Jenkins data found.")
        sys.exit(1)

    latest_date = sorted(results.keys(), reverse=True)[0]
    rows = []

    for env, data in results[latest_date].items():
        total = data["scenarios"]["total"]
        passed = data["scenarios"]["passed"]
        failed = data["scenarios"]["failed"]
        skipped = data["scenarios"]["skipped"]
        stability = (passed / total * 100) if total > 0 else 0
        failure_pct = (failed / total * 100) if total > 0 else 0

        rows.append({
            "date": latest_date,
            "builds": ", ".join(map(str, sorted(data["builds"]))),
            "environment": env,
            "stability": round(stability, 2),
            "failure_percentage": round(failure_pct, 2),
            "scenarios_total": total,
            "scenarios_passed": passed,
            "scenarios_failed": failed,
            "scenarios_skipped": skipped,
            "steps_total": data["steps"]["total"],
            "steps_passed": data["steps"]["passed"],
            "steps_failed": data["steps"]["failed"],
            "steps_skipped": data["steps"]["skipped"],
            "failed_scenarios": "; ".join(sorted(set(data["failed_scenarios"])))
        })

    df = pd.DataFrame(rows)

    # Save outputs
    df.to_csv("jenkins_summary.csv", index=False)
    df.to_html("jenkins_summary.html", index=False)
    save_to_excel(df)

    send_email_report(
        subject=f"📊 Jenkins Daily Report — {latest_date}",
        df=df,
        attachments=["jenkins_summary.csv", "jenkins_summary.html"]
    )
